import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl

class ContactBookGUI:
    def __init__(self, master):
        self.master = master
        self.master.title("Книга контактов")
        self.contacts = []
        self.current_index = tk.IntVar(value=0)

        self.load_contacts()

        self.name_label = tk.Label(master, text="Имя:")
        self.name_label.grid(row=0, column=0, sticky="w", padx=10, pady=5)
        self.name_var = tk.StringVar()
        self.name_entry = tk.Entry(master, textvariable=self.name_var)
        self.name_entry.grid(row=0, column=1, padx=10, pady=5)

        self.surname_label = tk.Label(master, text="Фамилия:")
        self.surname_label.grid(row=1, column=0, sticky="w", padx=10, pady=5)
        self.surname_var = tk.StringVar()
        self.surname_entry = tk.Entry(master, textvariable=self.surname_var)
        self.surname_entry.grid(row=1, column=1, padx=10, pady=5)

        self.phone_label = tk.Label(master, text="Телефон:")
        self.phone_label.grid(row=2, column=0, sticky="w", padx=10, pady=5)
        self.phone_var = tk.StringVar()
        self.phone_entry = tk.Entry(master, textvariable=self.phone_var)
        self.phone_entry.grid(row=2, column=1, padx=10, pady=5)

        self.address_label = tk.Label(master, text="Адрес:")
        self.address_label.grid(row=3, column=0, sticky="w", padx=10, pady=5)
        self.address_var = tk.StringVar()
        self.address_entry = tk.Entry(master, textvariable=self.address_var)
        self.address_entry.grid(row=3, column=1, padx=10, pady=5)

        self.add_button = tk.Button(master, text="Добавить контакт", command=self.add_contact)
        self.add_button.grid(row=4, column=0, columnspan=2, padx=10, pady=5)

        self.delete_button = tk.Button(master, text="Удалить контакт", command=self.delete_contact)
        self.delete_button.grid(row=5, column=0, columnspan=2, padx=10, pady=5)

        self.prev_button = tk.Button(master, text="Предыдущий контакт", command=self.prev_contact)
        self.prev_button.grid(row=6, column=0, columnspan=2, padx=10, pady=5)

        self.next_button = tk.Button(master, text="Следующий контакт", command=self.next_contact)
        self.next_button.grid(row=7, column=0, columnspan=2, padx=10, pady=5)

        self.import_button = tk.Button(master, text="Импорт контактов", command=self.import_contacts)
        self.import_button.grid(row=8, column=0, columnspan=2, padx=10, pady=5)

        self.export_button = tk.Button(master, text="Экспорт контактов", command=self.export_contacts)
        self.export_button.grid(row=9, column=0, columnspan=2, padx=10, pady=5)

        self.display_contact()

    def load_contacts(self):
        try:
            wb = openpyxl.load_workbook("contacts.xlsx")
            sheet = wb.active
            for row in sheet.iter_rows(values_only=True):
                self.contacts.append(row)
        except FileNotFoundError:
            print("Файл с контактами не найден. Создается новый.")

    def save_contacts(self):
        wb = openpyxl.Workbook()
        sheet = wb.active
        for contact in self.contacts:
            sheet.append(contact)
        wb.save("contacts.xlsx")

    def add_contact(self):
        name = self.name_var.get()
        surname = self.surname_var.get()
        phone = self.phone_var.get()
        address = self.address_var.get()
        self.contacts.append((name, surname, phone, address))
        self.save_contacts()
        messagebox.showinfo("Книга контактов", "Контакт успешно добавлен.")

    def delete_contact(self):
        del self.contacts[self.current_index.get()]
        self.save_contacts()
        messagebox.showinfo("Книга контактов", "Контакт успешно удален.")
        self.current_index.set(0)
        self.display_contact()

    def prev_contact(self):
        if self.current_index.get() > 0:
            self.current_index.set(self.current_index.get() - 1)
            self.display_contact()

    def next_contact(self):
        if self.current_index.get() < len(self.contacts) - 1:
            self.current_index.set(self.current_index.get() + 1)
            self.display_contact()

    def display_contact(self):
        contact = self.contacts[self.current_index.get()]
        self.name_var.set(contact[0])
        self.surname_var.set(contact[1])
        self.phone_var.set(contact[2])
        self.address_var.set(contact[3])

    def import_contacts(self):
        file_path = filedialog.askopenfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            try:
                wb = openpyxl.load_workbook(file_path)
                sheet = wb.active
                self.contacts.clear()
                for row in sheet.iter_rows(values_only=True):
                    self.contacts.append(row)
                self.current_index.set(0)
                self.display_contact()
                messagebox.showinfo("Книга контактов", "Контакты успешно импортированы.")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось импортировать контакты: {str(e)}")

    def export_contacts(self):
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            try:
                wb = openpyxl.Workbook()
                sheet = wb.active
                for contact in self.contacts:
                    sheet.append(contact)
                wb.save(file_path)
                messagebox.showinfo("Книга контактов", "Контакты успешно экспортированы.")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось экспортировать контакты: {str(e)}")

def main():
    root = tk.Tk()
    app = ContactBookGUI(root)
    root.mainloop()

if __name__ == "__main__":
    main()
